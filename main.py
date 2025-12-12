from os import listdir
from collections import defaultdict
from openpyxl import load_workbook

variants = defaultdict(lambda: defaultdict(dict))

for file in listdir("bin/results/"):
    _, variant_s, type_s, right = file.split("-")
    n_s = right.split(".")[0]
    txt_lines = open(f"bin/results/{file}").read().split("\n")

    variants[int(variant_s)][int(type_s)][int(n_s)] = [
        float(txt_lines[3].split(":")[1][:-6]), 
        float(txt_lines[4].split(":")[1][:-6]), 
        float(txt_lines[5].split(":")[1][:-6])
    ]

for variant, types in variants.items():
    wb = load_workbook(f"result-{variant}.xlsx")
    ws = wb["Лист1"]

    int_t, float_t, double_t = types[1], types[2], types[3]

    for n, (base, sse, avx) in int_t.items():
        row = 7 + n // 512
        
        ws.cell(row, 3, base)
        ws.cell(row, 4, sse)
        ws.cell(row, 5, avx)

    for n, (base, sse, avx) in float_t.items():
        row = 7 + n // 512
        
        ws.cell(row, 8, base)
        ws.cell(row, 9, sse)
        ws.cell(row, 10, avx)

    for n, (base, sse, avx) in double_t.items():
        row = 7 + n // 512
        
        ws.cell(row, 13, base)
        ws.cell(row, 14, sse)
        ws.cell(row, 15, avx)

    wb.save(f"result-{variant}.xlsx")
